from pathlib import Path
import re
import xml.etree.ElementTree as ET

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DATA_DIR = Path("datasets")
PCF_DIR = DATA_DIR / "PCF"
OUTPUT_DIR = Path("output")
OUTPUT_FILE = OUTPUT_DIR / "PCF自动复核结果.xlsx"


FLAG_MAP = {
    "0": "禁止",
    "1": "允许",
    "2": "必须",
    "3": "退补",
    "4": "沪市退补",
    "5": "深市退补",
    "禁止": "禁止",
    "允许": "允许",
    "必须": "必须",
    "退补": "退补",
}

BOARD_PARAM = {
    "上海主板": ("上海主板股票默认溢价比例(%)", "上海主板股票默认折价比例(%)"),
    "深圳主板": ("深圳主板默认溢价比例(%)", "深圳主板默认折价比例(%)"),
    "创业板": ("创业板默认溢价比例(%)", "创业板默认折价比例(%)"),
    "科创板": ("科创板默认溢价比例(%)", "科创板默认折价比例(%)"),
    "北交所": ("北交所主板默认溢价比例(%)", "北交所主板默认折价比例(%)"),
    "港市": ("港市股票默认溢价比例(%)", None),
}


def local_name(tag):
    return tag.split("}", 1)[-1] if "}" in tag else tag


def clean_code(value, width=6):
    if pd.isna(value):
        return ""
    text = str(value).strip().replace(".0", "")
    return text.zfill(width) if text.isdigit() and len(text) <= width else text


def norm_number(value):
    if pd.isna(value):
        return np.nan
    text = str(value).strip().replace(",", "").replace("\t", "")
    if text in {"", "-", "nan", "None"}:
        return np.nan
    return pd.to_numeric(text.rstrip("%"), errors="coerce")


def norm_rate(value):
    num = norm_number(value)
    if pd.isna(num):
        return np.nan
    text = str(value).strip()
    if text.endswith("%") or abs(num) > 1:
        return num / 100
    return num


def values_match(expected, actual, tolerance=1e-6, rate=False):
    left = norm_rate(expected) if rate else norm_number(expected)
    right = norm_rate(actual) if rate else norm_number(actual)
    if pd.isna(left) and pd.isna(right):
        return True
    if pd.isna(left) or pd.isna(right):
        return False
    return abs(left - right) <= tolerance


def rate_matches(expected, actual):
    expected_rate = norm_rate(expected)
    actual_rate = norm_rate(actual)
    if not pd.isna(expected_rate) and abs(expected_rate) <= 1e-12 and pd.isna(actual_rate):
        return True
    return values_match(expected, actual, rate=True)


def security_board(code, market):
    code = str(code).strip()
    market = str(market or "").strip()
    if market == "103":
        return "港市"
    if code.startswith(("688", "689")):
        return "科创板"
    if code.startswith(("300", "301", "302")):
        return "创业板"
    if code.startswith(("600", "601", "603", "605")):
        return "上海主板"
    if code.startswith(("000", "001", "002", "003")):
        return "深圳主板"
    if code.startswith(("430", "83", "87", "88", "92")):
        return "北交所"
    return "未识别"


def check_rows_product(checks, code, level, module, item, ok, expected, actual, note=""):
    status = "PASS" if ok else ("WARN" if level == "警告" else "FAIL")
    checks.append(
        {
            "产品代码": code,
            "严重级别": level,
            "模块": module,
            "校验项": item,
            "状态": status,
            "期望值": expected,
            "实际值": actual,
            "说明": note,
        }
    )


def load_inputs():
    params_long = pd.read_excel(DATA_DIR / "投资邮件.xlsx", sheet_name="参数信息", dtype={"产品代码": str})
    params_long["产品代码"] = params_long["产品代码"].map(clean_code)
    params_long["PCF日期"] = params_long["PCF日期"].astype(str).str.replace(r"\.0$", "", regex=True)
    params_long["参数名称"] = params_long["参数名称"].astype(str).str.strip()

    params_wide = (
        params_long.pivot_table(
            index=["产品代码", "产品名称", "PCF日期"],
            columns="参数名称",
            values="参数值",
            aggfunc="first",
        )
        .reset_index()
    )
    params_wide.columns.name = None

    valuation = pd.read_excel(DATA_DIR / "T-1日估值数据.xlsx", sheet_name=0, dtype=str)
    valuation = valuation.rename(
        columns={
            "产品": "组合",
            "产品名称": "基金名称",
            "代码": "基金代码",
            "总份额": "基金总份额",
            "总净值": "基金总净值",
            "单位净值": "基金单位净值",
        }
    )
    valuation = valuation.dropna(subset=["基金代码"]).copy()
    valuation["产品代码"] = valuation["基金代码"].map(clean_code)
    valuation["估值日期"] = pd.to_datetime(valuation["日期"], errors="coerce").dt.strftime("%Y%m%d")
    for col in ["基金总份额", "基金总净值", "基金单位净值", "创设单位净值", "现金差额"]:
        valuation[col] = pd.to_numeric(
            valuation[col]
            .astype(str)
            .str.strip()
            .str.replace(",", "", regex=False)
            .replace({"-": np.nan, "nan": np.nan, "None": np.nan}),
            errors="coerce",
        )
    valuation = valuation[
        [
            "估值日期",
            "产品代码",
            "组合",
            "基金名称",
            "基金总份额",
            "基金总净值",
            "基金单位净值",
            "创设单位净值",
            "现金差额",
        ]
    ]

    special = pd.read_excel(
        DATA_DIR / "投资邮件.xlsx",
        sheet_name="成份券特殊设置",
        dtype={"产品代码": str, "股票代码": str},
    )
    special = special.dropna(subset=["产品代码", "股票代码"]).copy()
    special["产品代码"] = special["产品代码"].map(clean_code)
    special["股票代码"] = special["股票代码"].map(clean_code)
    special["PCF日期"] = special["PCF日期"].astype(str).str.replace(r"\.0$", "", regex=True)
    special["替代标志_标准"] = (
        special["替代标志"].astype(str).str.strip().map(FLAG_MAP).fillna(special["替代标志"].astype(str).str.strip())
    )
    special["数量_数值"] = pd.to_numeric(
        special["数量"].astype(str).str.strip().replace({"-": np.nan, "nan": np.nan, "None": np.nan}),
        errors="coerce",
    )
    return params_wide, valuation, special


def parse_pcf_files():
    products = []
    components = []
    for path in sorted(PCF_DIR.glob("*.xml")):
        root = ET.parse(path).getroot()
        root_tag = local_name(root.tag)
        file_match = re.search(r"(\d{6})_(\d{8})", path.stem)
        file_code = file_match.group(1) if file_match else ""
        file_date = file_match.group(2) if file_match else ""
        top = {local_name(child.tag): (child.text or "").strip() for child in list(root)}

        product = {
            "文件名": path.name,
            "文件路径": str(path.relative_to(DATA_DIR.parent)),
            "文件产品代码": file_code,
            "文件日期": file_date,
            "XML类型": root_tag,
            "文件大小": path.stat().st_size,
        }

        if root_tag == "PCFFile":
            field_map = {
                "产品代码": "SecurityID",
                "产品名称": "Symbol",
                "基金公司": "FundManagementCompany",
                "跟踪指数": "UnderlyingSecurityID",
                "最小申赎单位": "CreationRedemptionUnit",
                "交易日": "TradingDay",
                "上一交易日": "PreTradingDay",
                "T日预估现金差额": "EstimateCashComponent",
                "现金替代比例上限": "MaxCashRatio",
                "是否公布IOPV": "Publish",
                "申购状态": "Creation",
                "赎回状态": "Redemption",
                "成份券记录数": "RecordNum",
                "总记录数": "TotalRecordNum",
                "现金差额": "CashComponent",
                "最小申赎单位净值": "NAVperCU",
                "基金份额净值": "NAV",
                "现金红利": "DividendPerCU",
                "累计申购总额限制": "CreationLimit",
                "累计赎回总额限制": "RedemptionLimit",
                "单账户累计申购限制": "CreationLimitPerUser",
                "单账户累计赎回限制": "RedemptionLimitPerUser",
            }
            for std_col, raw_col in field_map.items():
                product[std_col] = top.get(raw_col)
            comp_parent = next((x for x in root.iter() if local_name(x.tag) == "Components"), None)
            if comp_parent is not None:
                for comp in list(comp_parent):
                    if local_name(comp.tag) != "Component":
                        continue
                    row = {local_name(c.tag): (c.text or "").strip() for c in list(comp)}
                    code = clean_code(row.get("UnderlyingSecurityID", ""))
                    name = row.get("UnderlyingSymbol", "")
                    components.append(
                        {
                            "文件名": path.name,
                            "产品代码": clean_code(product.get("产品代码", "")),
                            "XML类型": root_tag,
                            "证券代码": code,
                            "证券名称": name,
                            "证券市场": row.get("UnderlyingSecurityIDSource"),
                            "数量": norm_number(row.get("ComponentShare")),
                            "替代标志原值": row.get("SubstituteFlag"),
                            "替代标志": FLAG_MAP.get(str(row.get("SubstituteFlag", "")).strip(), str(row.get("SubstituteFlag", "")).strip()),
                            "申购溢价比例": norm_rate(row.get("PremiumRatio")),
                            "赎回折价比例": norm_rate(row.get("DiscountRatio")),
                            "申购现金替代金额": norm_number(row.get("CreationCashSubstitute")),
                            "赎回现金替代金额": norm_number(row.get("RedemptionCashSubstitute")),
                            "是否现金记录": "现金" in str(name),
                        }
                    )

        elif root_tag == "ETFDefinition":
            field_map = {
                "产品代码": "FundInstrumentID",
                "产品名称": "FundName",
                "基金公司": "FundCompanyName",
                "跟踪指数": "UnderlyingIndex",
                "最小申赎单位": "CreationRedemptionUnit",
                "交易日": "TradingDay",
                "上一交易日": "PreTradingDay",
                "最小申赎单位净值": "NAVperCU",
                "基金份额净值": "NAV",
                "现金差额": "PreCashComponent",
                "现金红利": "CashDividend",
                "T日预估现金差额": "EstimatedCashComponent",
                "现金替代比例上限": "MaxCashRatio",
                "累计赎回总额限制": "RedemptionLimit",
                "申赎开关": "CreationRedemptionSwitch",
                "申赎机制": "CreationRedemptionMechanism",
                "成份券记录数": "RecordNumber",
            }
            for std_col, raw_col in field_map.items():
                product[std_col] = top.get(raw_col)
            comp_parent = next((x for x in root.iter() if local_name(x.tag) == "ComponentList"), None)
            if comp_parent is not None:
                for comp in list(comp_parent):
                    if local_name(comp.tag) != "Component":
                        continue
                    row = {local_name(c.tag): (c.text or "").strip() for c in list(comp)}
                    code = clean_code(row.get("InstrumentID", ""))
                    name = row.get("InstrumentName", "")
                    components.append(
                        {
                            "文件名": path.name,
                            "产品代码": clean_code(product.get("产品代码", "")),
                            "XML类型": root_tag,
                            "证券代码": code,
                            "证券名称": name,
                            "证券市场": row.get("UnderlyingSecurityID"),
                            "数量": norm_number(row.get("Quantity")),
                            "替代标志原值": row.get("SubstitutionFlag"),
                            "替代标志": FLAG_MAP.get(str(row.get("SubstitutionFlag", "")).strip(), str(row.get("SubstitutionFlag", "")).strip()),
                            "申购溢价比例": norm_rate(row.get("CreationPremiumRate")),
                            "赎回折价比例": norm_rate(row.get("RedemptionDiscountRate")),
                            "申购现金替代金额": norm_number(row.get("SubstitutionCashAmount")),
                            "赎回现金替代金额": np.nan,
                            "是否现金记录": "现金" in str(name),
                        }
                    )
        product["产品代码"] = clean_code(product.get("产品代码", ""))
        products.append(product)

    pcf_products = pd.DataFrame(products)
    pcf_components = pd.DataFrame(components)
    for col in [
        "最小申赎单位",
        "T日预估现金差额",
        "现金替代比例上限",
        "成份券记录数",
        "总记录数",
        "现金差额",
        "最小申赎单位净值",
        "基金份额净值",
        "累计申购总额限制",
        "累计赎回总额限制",
    ]:
        if col in pcf_products.columns:
            pcf_products[col] = pcf_products[col].map(norm_number)
    return pcf_products, pcf_components


def parse_flags():
    rows = []
    for path in sorted(list(PCF_DIR.glob("*.flag")) + list(PCF_DIR.glob("*.flg"))):
        raw = path.read_bytes()
        text = raw.decode("utf-8", errors="ignore")
        file_match = re.search(r"(\d{6})_(\d{8})", path.stem)
        row = {
            "flag文件名": path.name,
            "flag路径": str(path.relative_to(DATA_DIR.parent)),
            "文件产品代码": file_match.group(1) if file_match else "",
            "文件日期": file_match.group(2) if file_match else "",
            "flag大小": path.stat().st_size,
            "登记XML文件名": None,
            "登记XML大小": np.nan,
            "格式": "pipe_flg" if path.suffix.lower() == ".flg" else "xml_flag",
        }
        if path.suffix.lower() == ".flg":
            parts = [x.strip() for x in text.split("|")]
            row["登记XML文件名"] = parts[0] if parts else None
            row["登记XML大小"] = pd.to_numeric(parts[1], errors="coerce") if len(parts) > 1 else np.nan
        else:
            try:
                root = ET.fromstring(raw)
                flat = {local_name(e.tag): (e.text or "").strip() for e in root.iter()}
                row["登记XML文件名"] = flat.get("FileName")
                row["登记XML大小"] = pd.to_numeric(flat.get("FileSize"), errors="coerce")
            except Exception:
                match = re.search(r"(pcf_\d{6}_\d{8}\.xml|etfd_\d{6}_\d{8}\.xml)", text)
                row["登记XML文件名"] = match.group(1) if match else None
        rows.append(row)
    return pd.DataFrame(rows)


def build_checks(params_wide, valuation, special, pcf_products, pcf_components, flags):
    checks = []
    param_by_code = {r["产品代码"]: r for _, r in params_wide.iterrows()}
    pcf_by_code = {r["产品代码"]: r for _, r in pcf_products.iterrows()}
    valuation_by_code = {r["产品代码"]: r for _, r in valuation.iterrows()}
    all_codes = sorted(set(params_wide["产品代码"]) | set(pcf_products["产品代码"]) | set(valuation["产品代码"]))

    for code in all_codes:
        pcf = pcf_by_code.get(code)
        prm = param_by_code.get(code)
        val = valuation_by_code.get(code)

        if pcf is None:
            note = "估值表有该产品，但投资邮件/PCF目录未覆盖；如非当日复核范围，可人工确认后忽略"
            check_rows_product(checks, code, "错误", "文件完整性", "邮件/估值产品存在PCF文件", False, "存在", "缺失", note)
            continue
        if prm is None:
            check_rows_product(checks, code, "错误", "文件完整性", "PCF产品存在邮件参数", False, "存在", "缺失")
            continue

        check_rows_product(checks, code, "信息", "文件完整性", "邮件产品存在PCF文件", True, "存在", pcf["文件名"])
        check_rows_product(
            checks,
            code,
            "错误",
            "日期一致性",
            "邮件PCF日期=XML交易日",
            str(prm.get("PCF日期", "")) == str(pcf.get("交易日", "")).replace(".0", ""),
            prm.get("PCF日期"),
            pcf.get("交易日"),
        )
        check_rows_product(
            checks,
            code,
            "错误",
            "产品一致性",
            "文件名产品代码=XML产品代码",
            pcf.get("文件产品代码") == pcf.get("产品代码"),
            pcf.get("文件产品代码"),
            pcf.get("产品代码"),
        )

        for mail_col, pcf_col, label in [
            ("最小申购、赎回单位", "最小申赎单位", "最小申购赎回单位"),
            ("现金替代比例上限", "现金替代比例上限", "现金替代比例上限"),
            ("累计赎回总额限制", "累计赎回总额限制", "累计赎回总额限制"),
            ("累计申购总额限制", "累计申购总额限制", "累计申购总额限制"),
        ]:
            expected_raw = prm.get(mail_col, np.nan)
            actual_raw = pcf.get(pcf_col, np.nan)
            if str(expected_raw).strip() == "-":
                ok = pd.isna(norm_number(actual_raw)) or abs(norm_number(actual_raw)) <= 1e-6
            else:
                ok = values_match(expected_raw, actual_raw, rate=("比例" in label))
            check_rows_product(checks, code, "错误", "参数一致性", label, ok, expected_raw, actual_raw)

        if val is None:
            check_rows_product(checks, code, "警告", "T-1估值一致性", "PCF产品存在T-1估值数据", False, "存在", "缺失")
        else:
            check_rows_product(
                checks,
                code,
                "错误",
                "T-1估值一致性",
                "估值日期=XML上一交易日",
                str(val.get("估值日期", "")) == str(pcf.get("上一交易日", "")).replace(".0", ""),
                val.get("估值日期"),
                pcf.get("上一交易日"),
                val.get("基金名称", ""),
            )
            for val_col, pcf_col, label, tolerance in [
                ("基金单位净值", "基金份额净值", "基金单位净值", 1e-4),
                ("创设单位净值", "最小申赎单位净值", "创设单位净值", 0.01),
                ("现金差额", "现金差额", "现金差额", 0.01),
            ]:
                check_rows_product(
                    checks,
                    code,
                    "错误",
                    "T-1估值一致性",
                    label,
                    values_match(val.get(val_col, np.nan), pcf.get(pcf_col, np.nan), tolerance=tolerance),
                    val.get(val_col, np.nan),
                    pcf.get(pcf_col, np.nan),
                    val.get("基金名称", ""),
                )

        allow_text = str(prm.get("申购赎回的允许情况", "")).strip()
        if pcf.get("XML类型") == "PCFFile":
            actual_switch = f"申购={pcf.get('申购状态')};赎回={pcf.get('赎回状态')}"
            ok = "允许申购和赎回" in allow_text and pcf.get("申购状态") == "Y" and pcf.get("赎回状态") == "Y"
        else:
            actual_switch = pcf.get("申赎开关")
            ok = "允许申购和赎回" in allow_text and str(actual_switch) in {"1", "1.0"}
        check_rows_product(checks, code, "错误", "申赎状态", "申购赎回允许情况", ok, allow_text, actual_switch)

        comp_rows = pcf_components[pcf_components["产品代码"] == code]
        comp_count = len(comp_rows)
        if pcf.get("XML类型") == "PCFFile" and not pd.isna(pcf.get("总记录数", np.nan)):
            record_num = pcf.get("总记录数")
            item = "成份券明细行数=XML总记录数"
        else:
            record_num = pcf.get("成份券记录数")
            item = "成份券明细行数=XML记录数"
        check_rows_product(checks, code, "错误", "成份券完整性", item, values_match(record_num, comp_count), record_num, comp_count)

        dup = int(comp_rows["证券代码"].duplicated().sum())
        check_rows_product(checks, code, "错误", "成份券完整性", "成份券代码无重复", dup == 0, 0, dup)

    for _, flag in flags.iterrows():
        code = flag["文件产品代码"]
        xml_name = flag["登记XML文件名"]
        if not xml_name:
            xml_name = f"pcf_{code}_{flag['文件日期']}.xml" if flag["flag文件名"].startswith("pcf_") else f"etfd_{code}_{flag['文件日期']}.xml"
        xml_path = PCF_DIR / str(xml_name)
        check_rows_product(checks, code, "错误", "文件完整性", "flag登记XML存在", xml_path.exists(), xml_name, "存在" if xml_path.exists() else "缺失", flag["flag文件名"])
        if xml_path.exists() and not pd.isna(flag["登记XML大小"]):
            check_rows_product(
                checks,
                code,
                "错误",
                "文件完整性",
                "flag登记XML大小一致",
                int(flag["登记XML大小"]) == xml_path.stat().st_size,
                int(flag["登记XML大小"]),
                xml_path.stat().st_size,
                flag["flag文件名"],
            )

    comp_index = pcf_components.set_index(["产品代码", "证券代码"], drop=False)
    for _, row in special.iterrows():
        key = (row["产品代码"], row["股票代码"])
        if key not in comp_index.index:
            check_rows_product(checks, row["产品代码"], "错误", "特殊成份券", "特殊设置成份券存在于PCF", False, row["股票代码"], "缺失", row.get("股票简称", ""))
            continue
        comp = comp_index.loc[key]
        comp = comp.iloc[0] if isinstance(comp, pd.DataFrame) else comp
        expected_flag = row["替代标志_标准"]
        actual_flag = comp["替代标志"]
        check_rows_product(checks, row["产品代码"], "错误", "特殊成份券", f"{row['股票代码']} 替代标志", expected_flag == actual_flag, expected_flag, actual_flag, row.get("股票简称", ""))
        if not pd.isna(row["数量_数值"]):
            check_rows_product(checks, row["产品代码"], "错误", "特殊成份券", f"{row['股票代码']} 数量", values_match(row["数量_数值"], comp["数量"]), row["数量"], comp["数量"], row.get("股票简称", ""))

    fee_scope = pcf_components[
        (~pcf_components["是否现金记录"]) & (pcf_components["替代标志"] == "允许")
    ]
    for _, comp in fee_scope.iterrows():
        code = comp["产品代码"]
        prm = param_by_code.get(code)
        if prm is None:
            continue
        board = security_board(comp["证券代码"], comp["证券市场"])
        premium_col, discount_col = BOARD_PARAM.get(board, (None, None))
        if premium_col is None:
            check_rows_product(checks, code, "警告", "溢价折价比例", f"{comp['证券代码']} 板块识别", False, "可识别板块", board, comp.get("证券名称", ""))
            continue

        expected_premium = prm.get(premium_col, np.nan)
        if not pd.isna(norm_rate(expected_premium)):
            check_rows_product(
                checks,
                code,
                "错误",
                "溢价折价比例",
                f"{comp['证券代码']} 申购溢价比例",
                rate_matches(expected_premium, comp["申购溢价比例"]),
                expected_premium,
                comp["申购溢价比例"],
                f"{comp.get('证券名称', '')};{board}",
            )

        if discount_col:
            expected_discount = prm.get(discount_col, np.nan)
            if not pd.isna(norm_rate(expected_discount)):
                check_rows_product(
                    checks,
                    code,
                    "错误",
                    "溢价折价比例",
                    f"{comp['证券代码']} 赎回折价比例",
                    rate_matches(expected_discount, comp["赎回折价比例"]),
                    expected_discount,
                    comp["赎回折价比例"],
                    f"{comp.get('证券名称', '')};{board}",
                )

    checks_df = pd.DataFrame(checks)
    checks_df["是否异常"] = checks_df["状态"].isin(["FAIL", "WARN"])
    return checks_df


def write_output(params_wide, valuation, special, pcf_products, pcf_components, flags, checks_df):
    summary = (
        checks_df.groupby("产品代码")
        .agg(
            校验项数量=("校验项", "count"),
            错误数=("状态", lambda x: int((x == "FAIL").sum())),
            警告数=("状态", lambda x: int((x == "WARN").sum())),
        )
        .reset_index()
    )
    summary["复核结论"] = np.select([summary["错误数"] > 0, summary["警告数"] > 0], ["不通过", "需关注"], default="通过")
    summary = summary.merge(pcf_products[["产品代码", "产品名称", "XML类型", "文件名"]], on="产品代码", how="left")
    summary = summary[["产品代码", "产品名称", "XML类型", "文件名", "校验项数量", "错误数", "警告数", "复核结论"]]

    module_summary = (
        checks_df.groupby(["模块", "状态"])
        .size()
        .reset_index(name="数量")
        .sort_values(["模块", "状态"])
    )
    exception_only = checks_df[checks_df["是否异常"]].copy()

    OUTPUT_DIR.mkdir(exist_ok=True)
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="汇总", index=False)
        exception_only.to_excel(writer, sheet_name="异常明细", index=False)
        module_summary.to_excel(writer, sheet_name="模块统计", index=False)
        checks_df.to_excel(writer, sheet_name="全量明细", index=False)
        pcf_products.to_excel(writer, sheet_name="PCF产品层", index=False)
        pcf_components.to_excel(writer, sheet_name="PCF成份券", index=False)
        params_wide.to_excel(writer, sheet_name="邮件参数", index=False)
        valuation.to_excel(writer, sheet_name="T-1估值数据", index=False)
        special.to_excel(writer, sheet_name="特殊设置", index=False)
        flags.to_excel(writer, sheet_name="flag文件", index=False)

    wb = load_workbook(OUTPUT_FILE)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    bad_fill = PatternFill("solid", fgColor="FCE4D6")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    pass_fill = PatternFill("solid", fgColor="E2F0D9")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        headers = [cell.value for cell in ws[1]]
        status_col = headers.index("状态") + 1 if "状态" in headers else None
        conclusion_col = headers.index("复核结论") + 1 if "复核结论" in headers else None
        for row in range(2, ws.max_row + 1):
            marker = ws.cell(row=row, column=status_col).value if status_col else ws.cell(row=row, column=conclusion_col).value if conclusion_col else None
            fill = None
            if marker in {"FAIL", "不通过"}:
                fill = bad_fill
            elif marker in {"WARN", "需关注"}:
                fill = warn_fill
            elif marker in {"PASS", "通过"}:
                fill = pass_fill
            if fill:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, 200) + 1))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)
    wb.save(OUTPUT_FILE)
    return summary, module_summary, exception_only


def run_review(data_dir=DATA_DIR, output_dir=OUTPUT_DIR):
    global DATA_DIR, PCF_DIR, OUTPUT_DIR, OUTPUT_FILE

    old_data_dir = DATA_DIR
    old_pcf_dir = PCF_DIR
    old_output_dir = OUTPUT_DIR
    old_output_file = OUTPUT_FILE

    DATA_DIR = Path(data_dir)
    PCF_DIR = DATA_DIR / "PCF"
    OUTPUT_DIR = Path(output_dir)
    OUTPUT_FILE = OUTPUT_DIR / "PCF自动复核结果.xlsx"

    try:
        params_wide, valuation, special = load_inputs()
        pcf_products, pcf_components = parse_pcf_files()
        flags = parse_flags()
        checks_df = build_checks(params_wide, valuation, special, pcf_products, pcf_components, flags)
        summary, module_summary, exception_only = write_output(
            params_wide, valuation, special, pcf_products, pcf_components, flags, checks_df
        )

        return {
            "output_file": OUTPUT_FILE,
            "summary": summary,
            "module_summary": module_summary,
            "exception_only": exception_only,
            "checks": checks_df,
        }
    finally:
        DATA_DIR = old_data_dir
        PCF_DIR = old_pcf_dir
        OUTPUT_DIR = old_output_dir
        OUTPUT_FILE = old_output_file


def main():
    result = run_review()
    summary = result["summary"]
    module_summary = result["module_summary"]
    exception_only = result["exception_only"]
    checks_df = result["checks"]
    output_file = result["output_file"]

    print(f"输出文件: {output_file}")
    print(f"产品数: {len(summary)}")
    print(f"校验项: {len(checks_df)}")
    print(f"异常项: {len(exception_only)}")
    print(module_summary.to_string(index=False))


if __name__ == "__main__":
    main()
